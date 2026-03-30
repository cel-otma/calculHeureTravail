import pandas as pd
import re
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, json, os

# ── helpers ──────────────────────────────────────────────────────────────────

def parse_employee(raw):
    raw = str(raw).strip()
    m = re.match(r'(\d+)\((.+)\)', raw)
    if m:
        return m.group(1), m.group(2).strip()
    return raw, raw


def detect_entry_exit(equipement):
    eq = str(equipement).upper()
    if re.search(r'[_\-]E[\-\d]', eq):
        return 'E'
    if re.search(r'[_\-]S[\-\d]', eq):
        return 'S'
    m = re.search(r'[_\-]([ES])$', eq)
    if m:
        return m.group(1)
    return 'E'


def parse_date_value(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().split(' ')[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None


def parse_time_value(v):
    if isinstance(v, datetime):
        return v.time()
    s = str(v).strip()
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(s, fmt).time()
        except:
            pass
    try:
        frac = float(s)
        total_s = round(frac * 86400)
        h, rem = divmod(total_s, 3600)
        m, sec = divmod(rem, 60)
        return datetime.strptime(f'{h%24:02d}:{m:02d}:{sec:02d}', '%H:%M:%S').time()
    except:
        return None


def fmt_hhmm(td):
    if td is None:
        return ''
    total = int(td.total_seconds())
    if total < 0:
        return ''
    h, rem = divmod(total, 3600)
    return f'{h:02d}:{rem//60:02d}'


def time_to_str(t):
    if t is None:
        return ''
    return t.strftime('%H:%M:%S') if hasattr(t, 'strftime') else str(t)


# ── core ──────────────────────────────────────────────────────────────────────

def process_pointage(input_path: str, output_path: str) -> dict:
    try:
        # Read without dtype=str first so dates come as datetime objects
        all_sheets = pd.read_excel(input_path, sheet_name=None, header=0)
    except Exception as e:
        return {'success': False, 'error': f'Impossible de lire le fichier : {e}'}

    frames = []
    for sheet_name, df in all_sheets.items():
        df.columns = [str(c).strip() for c in df.columns]

        col_map = {}
        for col in df.columns:
            cl = col.lower().replace('é', 'e').replace('è', 'e').replace('ê', 'e')
            if 'date' in cl and 'col_date' not in col_map:
                col_map['col_date'] = col
            elif 'heure' in cl and 'col_heure' not in col_map:
                col_map['col_heure'] = col
            elif ('equipement' in cl or 'quipement' in cl) and 'col_equip' not in col_map:
                col_map['col_equip'] = col
            elif 'utilisateur' in cl and 'groupe' not in cl and 'col_user' not in col_map:
                col_map['col_user'] = col

        if any(k not in col_map for k in ('col_date', 'col_heure', 'col_equip', 'col_user')):
            continue

        sub = df[[col_map['col_date'], col_map['col_heure'],
                   col_map['col_equip'], col_map['col_user']]].copy()
        sub.columns = ['date_raw', 'heure_raw', 'equipement', 'utilisateur']
        frames.append(sub)

    if not frames:
        return {'success': False, 'error': 'Aucune feuille valide trouvée. Colonnes requises : Date, Heure, Equipement, Utilisateur.'}

    df = pd.concat(frames, ignore_index=True)
    df.dropna(subset=['date_raw', 'heure_raw', 'utilisateur'], inplace=True)

    df['date']      = df['date_raw'].apply(parse_date_value)
    df['heure']     = df['heure_raw'].apply(parse_time_value)
    df['type']      = df['equipement'].apply(detect_entry_exit)
    df['matricule'], df['nom'] = zip(*df['utilisateur'].astype(str).apply(parse_employee))

    df.dropna(subset=['date', 'heure'], inplace=True)
    df.sort_values(['matricule', 'date', 'heure'], inplace=True)

    if df.empty:
        return {'success': False, 'error': 'Aucune ligne valide après parsing des dates/heures.'}

    MAX_PAIRS = 5
    results = []

    for (mat, d), grp in df.groupby(['matricule', 'date']):
        nom = grp['nom'].iloc[0]
        entries = sorted(grp[grp['type'] == 'E']['heure'].tolist())
        exits   = sorted(grp[grp['type'] == 'S']['heure'].tolist())

        pairs, e_idx, s_idx = [], 0, 0
        while e_idx < len(entries):
            ent = entries[e_idx]
            while s_idx < len(exits) and exits[s_idx] <= ent:
                s_idx += 1
            srt = exits[s_idx] if s_idx < len(exits) else None
            if srt:
                s_idx += 1
            pairs.append((ent, srt))
            e_idx += 1

        row = {'Matricule': mat, 'Employé': nom, 'Date': d}
        total_td = timedelta()

        for i in range(MAX_PAIRS):
            if i < len(pairs):
                ent, srt = pairs[i]
                row[f'Entrée_{i+1}'] = time_to_str(ent)
                row[f'Sortie_{i+1}'] = time_to_str(srt) if srt else ''
                if srt:
                    diff = datetime.combine(datetime.today(), srt) - datetime.combine(datetime.today(), ent)
                    if diff.total_seconds() > 0:
                        total_td += diff
            else:
                row[f'Entrée_{i+1}'] = ''
                row[f'Sortie_{i+1}'] = ''

        row['Total'] = fmt_hhmm(total_td)
        row['_total_seconds'] = int(total_td.total_seconds())
        results.append(row)

    if not results:
        return {'success': False, 'error': 'Aucune donnée valide trouvée après traitement.'}

    res_df = pd.DataFrame(results).sort_values(['Matricule', 'Date'])

    # ── Excel output ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Pointage Résumé"

    def bdr(style='thin'):
        s = Side(border_style=style, color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def fill(color):
        return PatternFill('solid', fgColor=color)

    hdr_font  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    cell_font = Font(size=9, name='Arial')
    bold_font = Font(bold=True, size=9, name='Arial')
    center    = Alignment(horizontal='center', vertical='center')
    left_al   = Alignment(horizontal='left',   vertical='center')

    static   = ['Matricule', 'Employé', 'Date']
    pair_hdrs = []
    for i in range(1, MAX_PAIRS + 1):
        pair_hdrs += [f'Entrée {i}', f'Sortie {i}']
    all_hdrs = static + pair_hdrs + ['Total']

    ws.append(all_hdrs)
    ws.row_dimensions[1].height = 22

    for ci, hdr in enumerate(all_hdrs, 1):
        cell = ws.cell(1, ci, hdr)
        cell.font = hdr_font; cell.alignment = center; cell.border = bdr()
        if hdr in ('Matricule', 'Employé', 'Date'):
            cell.fill = fill('1B2A4A')
        elif hdr == 'Total':
            cell.fill = fill('4C1D95')
        else:
            is_entry = (ci - len(static) - 1) % 2 == 0
            cell.fill = fill('166534' if is_entry else '7F1D1D')

    for ri, (_, row) in enumerate(res_df.iterrows(), 2):
        bg = 'F0F4FF' if ri % 2 == 0 else 'FFFFFF'
        ws.cell(ri, 1, row['Matricule'])
        ws.cell(ri, 2, row['Employé'])
        ws.cell(ri, 3, str(row['Date']))
        for c in range(1, 4):
            cell = ws.cell(ri, c)
            cell.font = bold_font if c <= 2 else cell_font
            cell.alignment = center if c == 1 else left_al
            cell.border = bdr()
            cell.fill = fill('EFF6FF' if c == 1 else bg)

        for i in range(MAX_PAIRS):
            ev = row.get(f'Entrée_{i+1}', '')
            sv = row.get(f'Sortie_{i+1}', '')
            ce = ws.cell(ri, 4 + i*2,     ev)
            cs = ws.cell(ri, 4 + i*2 + 1, sv if sv else ('—' if ev else ''))
            for c in (ce, cs):
                c.font = cell_font; c.alignment = center; c.border = bdr()
            if ev:
                ce.fill = fill('ECFDF5')
                ce.font = Font(size=9, name='Arial', color='065F46', bold=True)
            if sv:
                cs.fill = fill('FFF1F2')
                cs.font = Font(size=9, name='Arial', color='9F1239', bold=True)
            elif ev:
                cs.fill = fill('FEF9C3')
                cs.font = Font(size=9, name='Arial', color='92400E')

        tc = ws.cell(ri, 4 + MAX_PAIRS*2, row['Total'])
        tc.font = Font(bold=True, size=9, name='Arial', color='4C1D95')
        tc.fill = fill('F5F3FF'); tc.alignment = center; tc.border = bdr('medium')
        ws.row_dimensions[ri].height = 18

    # Summary
    ws2 = wb.create_sheet("Récapitulatif par Employé")
    ws2.append(['Matricule', 'Employé', 'Nb Jours', 'Total Heures'])
    ws2.row_dimensions[1].height = 20
    for c in range(1, 5):
        cell = ws2.cell(1, c)
        cell.font = hdr_font; cell.fill = fill('1B2A4A')
        cell.alignment = center; cell.border = bdr()

    summary = (res_df.groupby(['Matricule', 'Employé'])
               .agg(nb_jours=('Date', 'count'), total_sec=('_total_seconds', 'sum'))
               .reset_index())
    for ri, (_, row) in enumerate(summary.iterrows(), 2):
        td = timedelta(seconds=int(row['total_sec']))
        ws2.cell(ri, 1, row['Matricule']); ws2.cell(ri, 2, row['Employé'])
        ws2.cell(ri, 3, int(row['nb_jours'])); ws2.cell(ri, 4, fmt_hhmm(td))
        for c in range(1, 5):
            cell = ws2.cell(ri, c)
            cell.font = cell_font; cell.alignment = center; cell.border = bdr()
            if ri % 2 == 0:
                cell.fill = fill('F0F4FF')

    # Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    for i in range(MAX_PAIRS):
        ws.column_dimensions[get_column_letter(4 + i*2)].width = 11
        ws.column_dimensions[get_column_letter(5 + i*2)].width = 11
    ws.column_dimensions[get_column_letter(4 + MAX_PAIRS*2)].width = 10
    for col, w in zip('ABCD', [12, 24, 10, 14]):
        ws2.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    ws2.freeze_panes = 'A2'
    wb.save(output_path)

    return {'success': True, 'rows': len(results), 'employees': res_df['Matricule'].nunique(), 'output': output_path}


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'success': False, 'error': 'Usage: process_pointage.py <input> <output>'}))
        sys.exit(1)
    print(json.dumps(process_pointage(sys.argv[1], sys.argv[2])))
